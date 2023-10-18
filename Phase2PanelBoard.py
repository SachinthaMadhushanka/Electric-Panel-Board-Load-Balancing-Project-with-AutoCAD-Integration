from itertools import combinations
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
import openpyxl
import time


def balance_diff(phase_power):
    max_power = max(phase_power.values())
    min_power = min(phase_power.values())
    return (max_power - min_power) * 100 / max_power


def canFitInPanelBoard(phase_power, max_spaces):
    max_length = max(len(lst) for lst in phase_power.values())

    spaces = 0
    for lst in phase_power.values():
        if len(lst) == max_length:
            continue
        spaces += max_length - len(lst) - 1

    return spaces <= max_spaces


def assign_circuits(circuits, index, phases, phase_power, best_phases, best_diff, memo, max_diff=4, max_spaces=22):
    if index == len(circuits):

        if not canFitInPanelBoard(dict(phases), max_spaces):
            return best_phases, best_diff

        # print(dict(phases))
        current_diff = balance_diff(phase_power)
        # print(current_diff)

        if current_diff < best_diff:
            return dict(phases), current_diff
        else:
            return best_phases, best_diff

    memo_key = (index, tuple(phase_power.values()))
    if memo_key in memo:
        return memo[memo_key]

    row_index, circuit_power, circuit_phases = circuits[index]
    divided_power = circuit_power / circuit_phases

    for phase_combination in combinations(['A', 'B'], circuit_phases):
        temp_phases = {k: list(v) for k, v in phases.items()}
        temp_phase_power = dict(phase_power)

        for phase in phase_combination:
            temp_phases[phase].append((row_index, divided_power))
            temp_phase_power[phase] += divided_power

        if balance_diff(temp_phase_power) <= max_diff:
            result_phases, result_diff = assign_circuits(circuits, index + 1, temp_phases, temp_phase_power,
                                                         temp_phases, best_diff, memo, max_diff, max_spaces)
            if result_diff < best_diff:
                best_phases = result_phases
                best_diff = result_diff
                if best_diff <= max_diff:
                    memo[memo_key] = best_phases, best_diff
                    return best_phases, best_diff
        else:
            result_phases, result_diff = assign_circuits(circuits, index + 1, temp_phases, temp_phase_power,
                                                         best_phases, best_diff, memo, max_diff, max_spaces)
            if result_diff < best_diff:
                best_phases = result_phases
                best_diff = result_diff

    memo[memo_key] = best_phases, best_diff
    return best_phases, best_diff


def distribute_phases(circuits, max_spaces):
    start_time = time.time()
    print("Start Distributing Among 2 Phases")

    circuits = sorted(circuits, key=lambda x: (-x[1], x[2]))
    phases = {'A': [], 'B': []}
    phase_power = {'A': 0, 'B': 0}
    memo = {}

    best_phases, best_diff = assign_circuits(circuits, 0, phases, phase_power, {}, float('inf'), memo, 4, max_spaces)

    print(
        f"Completed Distributing Among the Phases (Time Taken: {(time.time() - start_time)}, Difference: {best_diff})")
    return best_phases, best_diff


def getSpaces(best_phases):
    max_length = max(len(lst) for lst in best_phases.values())

    spaces = 0
    for lst in best_phases.values():
        if len(lst) == max_length:
            continue
        spaces += max_length - len(lst) - 1

    return spaces


def mergeCells(sheet, cell1, cell2):
    start_cell = f"{get_column_letter(cell1[1])}{cell1[0]}"
    end_cell = f"{get_column_letter(cell2[1])}{cell2[0]}"
    sheet.merge_cells(f"{start_cell}:{end_cell}")


"""
    Return sheet, total_power_list, excel_data
"""


def readDataFromFile(input_filename):
    print("\n-----------------------------------------------------------------------------------------")
    print(f"Start Reading From {input_filename}")
    work_book = openpyxl.load_workbook(input_filename, data_only=True)
    sheet = work_book.active

    total_power_col = sheet.max_column - 1
    poles_col = sheet.max_column

    """
        Read total powers into a list from Excel file
        total_power_list = [(3, 640, 1), (4, 345, 1), (5, 491, 1), (6, 225, 1), (7, 648, 1), (8, 810, 1), (9, 2000, 1), (10, 810, 1), (11, 986, 1), (12, 324, 1), (13, 700, 1), (14, 986, 1), (15, 824, 1), (16, 800, 1), (17, 162, 1), (18, 824, 1)]
    """

    total_power_list = []
    for row in range(3, sheet.max_row + 1):
        total_power_list.append((row, sheet.cell(row=row, column=total_power_col).value, sheet.cell(row=row, column=poles_col).value))

    """
        excel_data = [['CIRCUIT #1', None, 9, 3, 1, None, 11, 1, None, None, None, None, None, None, None, None, None, None, None, None, None, 640], ['CIRCUIT #2', 3, None, None, None, 5, None, None, None, 3, None, None, None, None, None, None, None, None, None, None, None, 345], . . . .]
    """
    excel_data = []
    for row in range(3, sheet.max_row + 1):
        # row_data = [sheet.cell(row=row, column=col) for col in range(1, sheet.max_column+1)]
        temp = []
        for col in range(1, sheet.max_column + 1):
            temp.append(sheet.cell(row=row, column=col).value)

        excel_data.append(temp)

    print("Completed Reading")
    return work_book, total_power_list, excel_data


# def generateOutputFile(work_book, best_phases, excel_data, output_filename):
#     """
#         Add first column to end
#     """
#
#     print("Generating Output File")
#     sheet = work_book.active
#     last_col_index = sheet.max_column
#     last_row_index = sheet.max_row
#
#     sheet.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
#     sheet.delete_cols(1)
#
#     """
#         Set headers of the Phase A, Phase B, Phase C columns
#     """
#     sheet.insert_cols(1, 2)
#     mergeCells(sheet, (1, 1), (1, 2))
#     sheet.cell(row=1, column=1).value = "Load Distribution"
#     sheet.cell(row=2, column=1).value = "Phase A"
#     sheet.cell(row=2, column=2).value = "Phase B"
#
#     """
#         Write balanced power to the excel sheet
#
#         all_lists = [[(9, 2000.0), (15, 824.0), (13, 700.0), (12, 324.0)], [(11, 986.0), (14, 986.0), (18, 824.0), (5, 491.0), (4, 345.0), (6, 225.0)], [(8, 810.0), (10, 810.0), (16, 800.0), (7, 648.0), (3, 640.0), (17, 162.0)]]
#
#         sorted_phase_list = [[(11, 986.0), (14, 986.0), (18, 824.0), (5, 491.0), (4, 345.0), (6, 225.0)], [(8, 810.0), (10, 810.0), (16, 800.0), (7, 648.0), (3, 640.0), (17, 162.0)], [(9, 2000.0), (15, 824.0), (13, 700.0), (12, 324.0)]]
#
#         sorted_phase_list => all_lists is sorted by list length in descending order
#     """
#     all_lists = [value for value in best_phases.values()]
#     sorted_phase_list = sorted(all_lists, key=lambda x: len(x), reverse=True)
#
#     # Set header to Circuit List cell
#     mergeCells(sheet, (1, last_col_index + 2), (2, last_col_index + 2))
#     sheet.cell(row=1, column=last_col_index + 2).value = "CIRCUIT LIST"
#
#     for sequence_num in range(len(sorted_phase_list[0])):
#
#         # For each Phase (Phase A, B)
#         for list_index in range(2):
#             table_index = sequence_num * 2 + list_index + 3
#
#             # For space circuit
#             if len(sorted_phase_list[list_index]) <= sequence_num:
#                 # Set all values to empty in that row
#                 for col in range(last_col_index + 2):
#                     sheet.cell(row=table_index, column=col + 1).value = ""
#                 continue
#
#             index_of_next_circuit = sorted_phase_list[list_index][sequence_num][0]
#
#             # Set load power data
#             for col in range(1, last_col_index):
#
#                 new_data = excel_data[index_of_next_circuit - 3][col]
#                 if not new_data:
#                     new_data = ""
#                 sheet.cell(row=table_index, column=col + 2).value = new_data
#
#             # Set Circuit name data at the end
#             sheet.cell(row=table_index, column=last_col_index + 2).value = excel_data[index_of_next_circuit - 3][0]
#
#             # Set total power column (In one of the first 3 cols)
#             sheet.cell(row=table_index, column=list_index + 1).value = sorted_phase_list[list_index][sequence_num][1]
#
#     """
#         Adding Cell Styles
#     """
#     total_valid_rows = last_row_index + getSpaces(best_phases)
#
#     thin_border = Border(left=Side(style='thin'),
#                          right=Side(style='thin'),
#                          top=Side(style='thin'),
#                          bottom=Side(style='thin'))
#
#     for row in range(1, total_valid_rows + 1):
#         for col in range(1, sheet.max_column + 1):
#             sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
#             sheet.cell(row=row, column=col).border = thin_border
#             sheet.cell(row=row, column=col).border = thin_border
#             sheet.cell(row=row, column=col).border = thin_border
#             sheet.cell(row=row, column=col).font = Font(size=8)
#
#     # Add fixed width to each column
#     for col in range(1, last_col_index + 3):
#         column_letter = get_column_letter(col)
#
#         width = 15
#         if col == last_col_index + 2:
#             width = 25
#
#         sheet.column_dimensions[column_letter].width = width
#
#     """
#         Highlighting
#     """
#     green = '00ff00'
#     yellow = 'ffff00'
#
#     green_highlight = PatternFill(start_color=green, end_color=green, fill_type='solid')
#     yellow_highlight = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
#
#     # Green Highlighting
#     for row_index in range(1, 3):
#         for row in sheet[get_column_letter(row_index)][:total_valid_rows]:
#             row.fill = green_highlight
#
#     # Yellow Highlighting
#     for row in sheet[get_column_letter(sheet.max_column)][:total_valid_rows]:
#         row.fill = yellow_highlight
#
#     work_book.save(output_filename)
#     print(f"Output File {output_filename} Generated")


def getNext2PoleLoadIndex(pole_2_lists, next_phase_start_index):
    minimumGapsLoadIndex = -1
    minimumGaps = 4

    for load_index in range(len(pole_2_lists[0])):
        phase_BC = pole_2_lists[0][load_index][1] == 0
        phase_CA = pole_2_lists[1][load_index][1] == 0

        # Get starting phase index
        # For phase_AB => 1, phase_BC = 2, phase_CA => 3
        cur_start_index = 1
        if phase_BC:
            cur_start_index = 2

        elif phase_CA:
            cur_start_index = 3

        # Calculate number of needed gaps for the current load
        gaps = cur_start_index - next_phase_start_index
        if gaps < 0:
            gaps += 3

        # Get minimum load
        if gaps < minimumGaps:
            minimumGaps = gaps
            minimumGapsLoadIndex = load_index

        # If there is no gaps
        if minimumGaps == 0:
            return load_index, 0

    return minimumGapsLoadIndex, minimumGaps


def WriteRow(sheet, row_id, is_empty, last_col_index=0, excel_data=None, index_excel=0, phase_power=0, phase_col=1):
    # Empty Row
    if is_empty:
        # Set all values to empty in that row
        for col in range(sheet.max_column):
            sheet.cell(row=row_id, column=col + 1).value = ""

    # Insert Pole 1 load in Phase A to the current row (Then Delete that load from pole_1_lists)
    else:
        # index_of_next_circuit = pole_1_lists[0][0][0]

        # Set load power data
        for col in range(1, last_col_index):

            new_data = excel_data[index_excel][col]
            if not new_data:
                new_data = ""
            sheet.cell(row=row_id, column=col + 2).value = new_data

        # Set Circuit name data at the end
        sheet.cell(row=row_id, column=last_col_index + 2).value = excel_data[index_excel][0]

        # Set Phase Power (In one of the first 3 cols)
        sheet.cell(row=row_id, column=phase_col).value = phase_power


def generateOutputFile(work_book, pole_1_lists, pole_2_lists, excel_data, output_filename):
    """
        Add first column to end
    """

    print("Generating Output File")
    sheet = work_book.active
    last_col_index = sheet.max_column
    last_row_index = sheet.max_row

    sheet.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.unmerge_cells(start_row=1, start_column=last_col_index, end_row=2, end_column=last_col_index)
    sheet.delete_cols(1)

    """
        Set headers of the Phase A, Phase B, Phase C columns
    """
    sheet.insert_cols(1, 2)
    mergeCells(sheet, (1, 1), (1, 2))
    sheet.cell(row=1, column=1).value = "Load Distribution"
    sheet.cell(row=2, column=1).value = "Phase A"
    sheet.cell(row=2, column=2).value = "Phase B"

    # Set header to Circuit List cell
    mergeCells(sheet, (1, last_col_index + 2), (2, last_col_index + 2))
    sheet.cell(row=1, column=last_col_index + 2).value = "CIRCUIT LIST"
    mergeCells(sheet, (1, last_col_index + 1), (2, last_col_index + 1))

    # """
    #     Write 3 Poles Loads
    # """
    #
    table_index = 2
    # Write pole 3 circuits data to the output file
    for sequence_num in range(len(pole_2_lists[0])):

        # Row location of the table
        table_index = sequence_num * 2 + 3

        # Get the index of the next circuit (To access excel data)
        index_of_next_circuit = pole_2_lists[0][sequence_num][0]

        # Merge 3 rows for load columns, total_power_columns and circuit_name column
        for col in range(3, sheet.max_column + 1):
            mergeCells(sheet, (sequence_num * 2 + 3, col), (sequence_num * 2 + 4, col))

        # Set circuit name
        sheet.cell(row=table_index, column=last_col_index + 2).value = excel_data[index_of_next_circuit - 3][0]

        # Set load power data
        for col in range(1, last_col_index):

            new_data = excel_data[index_of_next_circuit - 3][col]
            if not new_data:
                new_data = ""
            sheet.cell(row=table_index, column=col + 2).value = new_data

        # Set Phase Power
        # For each Phase (Phase A, B, C)
        for list_index in range(2):
            table_index = sequence_num * 2 + list_index + 3

            # Set total power column (In one of the 3 Phases)
            sheet.cell(row=table_index, column=list_index + 1).value = pole_2_lists[list_index][sequence_num][1]

    """
        Write 1 Pole Loads
    """

    table_index += 1
    next_phase_start_index = 1
    while len(pole_1_lists[0]) != 0 or len(pole_1_lists[1]) != 0:

        # If there is no 1 Pole load for that phase
        if len(pole_1_lists[next_phase_start_index - 1]) == 0:
            WriteRow(sheet, table_index, True)


        # Insert Pole 1 load in that Phase to the current row (Then Delete that load from pole_1_lists)
        else:
            index_of_next_circuit = pole_1_lists[next_phase_start_index - 1][0][0]

            WriteRow(sheet=sheet, row_id=table_index, is_empty=False, last_col_index=last_col_index,
                     excel_data=excel_data, index_excel=index_of_next_circuit - 3,
                     phase_power=pole_1_lists[next_phase_start_index - 1][0][1], phase_col=next_phase_start_index)

            # Deleting Already assigned load
            del pole_1_lists[next_phase_start_index - 1][0]

        table_index += 1
        if next_phase_start_index == 1:
            next_phase_start_index = 2
        else:
            next_phase_start_index = 1

    """
        Adding Cell Styles
    """
    total_valid_rows = table_index

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in range(1, total_valid_rows):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).font = Font(size=8)

    # Add fixed width to each column
    for col in range(1, last_col_index + 4):
        column_letter = get_column_letter(col)

        width = 15
        if col == last_col_index + 3:
            width = 25

        sheet.column_dimensions[column_letter].width = width

    """
        Highlighting
    """
    green = '00ff00'
    yellow = 'ffff00'

    green_highlight = PatternFill(start_color=green, end_color=green, fill_type='solid')
    yellow_highlight = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

    # Green Highlighting
    for row_index in range(1, 3):
        for row in sheet[get_column_letter(row_index)][:total_valid_rows]:
            row.fill = green_highlight

    # Yellow Highlighting
    for row in sheet[get_column_letter(sheet.max_column)][:total_valid_rows]:
        row.fill = yellow_highlight

    work_book.save(output_filename)
    print(f"Output File {output_filename} Generated")


def balance2PhasePanelBoard(input_filename, output_filename, max_circuits):
    work_book, total_power_list, excel_data = readDataFromFile(input_filename)

    spaces = max_circuits - len(total_power_list)

    if spaces < 0:
        print("Maximum circuits should be equal or greater than the number of circuits")
        return

    best_phases, best_diff = distribute_phases(total_power_list, spaces)

    all_lists = [value for value in best_phases.values()]
    sorted_phase_list = sorted(all_lists, key=lambda x: len(x), reverse=True)

    # Divide each circuit according to their poles
    pole_2_lists = [[], [], []]
    pole_1_lists = [[], [], []]

    phase_index = 0
    for lst in sorted_phase_list:
        for circuit_index, power in lst:
            poles = [circuit_data[0] for lst in sorted_phase_list for circuit_data in lst].count(circuit_index)
            if poles == 1:
                pole_1_lists[phase_index].append((circuit_index, power))

            else:
                pole_2_lists[phase_index].append((circuit_index, power))

        phase_index += 1

    # generateOutputFile(work_book, best_phases, excel_data, output_filename)
    generateOutputFile(work_book, pole_1_lists, pole_2_lists, excel_data, output_filename)

    return best_phases

# TR_input_filename = "Circuit List - TR.xlsx"
# TR_output_filename = "TR.xlsx"
# max_circuits = 20
# best_phases_TR = balance2PhasePanelBoard(TR_input_filename, TR_output_filename, max_circuits)
#
# print(best_phases_TR)
