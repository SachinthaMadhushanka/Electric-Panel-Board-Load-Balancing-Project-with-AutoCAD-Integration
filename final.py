import time

import openpyxl
import pywintypes

from Phase2PanelBoard import balance2PhasePanelBoard
from Phase3PanelBoard import balance3PhasePanelBoard
from tkinter import Tk, Label, Entry, filedialog, Button, Frame

from openpyxl.utils import get_column_letter
from collections import Counter
from openpyxl.styles import Border, Side, Alignment, Font

from openpyxl import Workbook
import win32com.client

f = open("env.txt")
lines = f.readlines()
generated_input_filename = lines[0].split("=>")[1].strip()
balanced_output_filename = lines[1].split("=>")[1].strip()
autocad_output_filename = lines[2].split("=>")[1].strip()
f.close()


circuit_load_handles_dict = {}


def mergeCells(sheet, cell1, cell2):
    start_cell = f"{get_column_letter(cell1[1])}{cell1[0]}"
    end_cell = f"{get_column_letter(cell2[1])}{cell2[0]}"
    sheet.merge_cells(f"{start_cell}:{end_cell}")


def readInputFileFromAutocad(filename):
    global circuit_load_handles_dict
    circuit_load_handles_dict = {}

    print("Start Reading generated file from AutoCAD")
    load_details_dict = {}
    circuit_details_dict = {}
    circuit_poles = {}

    f = open(filename, 'r')

    lines = f.readlines()

    # Read panel board name, number of phases and number of circuits from autocad file
    panel_board_name = lines[0].split(":")[1].strip()
    num_of_phases = int(lines[1].split(":")[1].strip())
    max_circuits = int(lines[2].split(":")[1].strip())

    # Read line by line
    for line in lines[3:]:
        line = line.strip()

        data = line.split()
        circuit_number = data[0].split(":")[1].strip()
        object_id = data[1].split(":")[1].strip()
        poles = int(data[2].split(":")[1].strip())
        power = float(data[3].split(":")[1].strip())
        load_number = int(data[5].split(":")[1].strip())

        # Update circuit poles value
        circuit_poles[circuit_number] = poles

        # Add each unique load to load_details_dict
        # Each load can be accessed by load number

        if load_number not in load_details_dict.keys():
            load_details_dict[load_number] = {"object_id": object_id, "poles": poles, "power": power}

        # Update circuit details dict
        temp = []
        if circuit_number in circuit_details_dict.keys():
            temp = circuit_details_dict[circuit_number]
        temp.append(load_number)
        circuit_details_dict[circuit_number] = temp

        # Update circuit load handles details
        temp = []
        if circuit_number in circuit_load_handles_dict.keys():
            temp = circuit_load_handles_dict[circuit_number]
        temp.append(object_id)
        circuit_load_handles_dict[circuit_number] = temp

    sorted_load_numbers = sorted(load_details_dict.keys())
    for circuit_num in circuit_details_dict.keys():
        counter = Counter(circuit_details_dict[circuit_num])
        circuit_details_dict[circuit_num] = counter

    print("Read generated file from AutoCAD")
    return load_details_dict, sorted_load_numbers, circuit_details_dict, panel_board_name, num_of_phases, max_circuits, circuit_poles


def startBalancing():
    global circuit_load_handles_dict

    load_details_dict, sorted_load_numbers, circuit_details_dict, panel_board_name, num_of_phases, max_circuits, circuit_poles = readInputFileFromAutocad(
        autocad_output_filename)

    print("Start generating input excel file")
    workbook = Workbook()
    worksheet = workbook.active

    mergeCells(worksheet, (1, 1), (2, 1))
    worksheet.cell(row=1, column=1).value = "CIRCUIT LIST"

    for index, load_num in enumerate(sorted_load_numbers):
        worksheet.cell(row=1, column=index + 2).value = "LOAD #" + str(load_num)
        worksheet.cell(row=2, column=index + 2).value = load_details_dict[load_num]["power"]

    worksheet.cell(row=1, column=len(sorted_load_numbers) + 2).value = "Total power"
    worksheet.cell(row=2, column=len(sorted_load_numbers) + 2).value = "(WATTS)"
    mergeCells(worksheet, (1, len(sorted_load_numbers) + 3), (2, len(sorted_load_numbers) + 3))
    worksheet.cell(row=1, column=len(sorted_load_numbers) + 3).value = "Poles"

    for index, circuit_num in enumerate(circuit_details_dict.keys()):
        worksheet.cell(row=3 + index, column=1).value = "CIRCUIT #" + str(circuit_num)
        loads = circuit_details_dict[circuit_num]

        circuit_power = 0
        for load_num, count in loads.items():
            load_index = sorted_load_numbers.index(load_num)
            worksheet.cell(row=3 + index, column=2 + load_index).value = count
            circuit_power += load_details_dict[load_num]["power"] * count

        worksheet.cell(row=3 + index, column=2 + len(sorted_load_numbers)).value = circuit_power

        worksheet.cell(row=3 + index, column=1).value = "CIRCUIT #" + str(circuit_num)
        worksheet.cell(row=3 + index, column=len(sorted_load_numbers) + 3).value = circuit_poles[circuit_num]

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in range(1, len(circuit_details_dict.keys()) + 3):
        for col in range(1, len(sorted_load_numbers) + 4):
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).font = Font(size=10)

    # Add fixed width to each column
    for col in range(1, len(circuit_details_dict.keys()) + 4):
        width = 15
        if col == 1:
            width = 25
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = width

    workbook.save(generated_input_filename)
    print("Completed generating input excel file")

    if num_of_phases == 3:
        balance3PhasePanelBoard(generated_input_filename, balanced_output_filename, max_circuits)

    else:
        balance2PhasePanelBoard(generated_input_filename, balanced_output_filename, max_circuits)

    updateCircuitNumbers()


def is_file_open(file_path):
    try:
        # Try to open the file in append mode
        with open(file_path, 'a'):
            pass
    except IOError:
        # If an IOError is raised, the file is likely open in another process
        return True
    return False


def updateCircuitNumbers():
    # Update circuit numbers in ascending order in balanced excel file
    pre_order_circuits = []
    work_book = openpyxl.load_workbook(balanced_output_filename, data_only=True)
    sheet = work_book.active

    circuit_index_ascending = 1
    for row in range(3, sheet.max_row + 1):
        value = sheet.cell(row=row, column=sheet.max_column).value
        if value:
            pre_order_circuits.append(value.split("#")[1].strip())
            sheet.cell(row=row, column=sheet.max_column).value = "Circuit #" + str(circuit_index_ascending)
            circuit_index_ascending += 1

    work_book.save(balanced_output_filename)

    # Update autocad file circuit numbers

    all_handles_with_correct_circuit_num = {}
    for index, pre_order_circuit in enumerate(pre_order_circuits):
        cur_handles = circuit_load_handles_dict[pre_order_circuit]
        for handle in cur_handles:
            all_handles_with_correct_circuit_num[handle] = index + 1


    autocad_file_path = autocad_file_entry.get()
    while True:
        if is_file_open(autocad_file_path):
            print("AutoCAD file is opened by some other program.")
            time.sleep(1)

        else:
            print("AutoCAD file can be updated.")
            break

    # Open a .dwg file

    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.Documents.Open(autocad_file_path)  # Replace with your actual file path

    # Iterate over all entities in the model space
    for entity in doc.ModelSpace:
        if entity.Handle in all_handles_with_correct_circuit_num.keys():
            # If the entity is a BlockReference, print its attributes
            if entity.ObjectName == 'AcDbBlockReference':
                for attrib in entity.GetAttributes():
                    print(attrib.TextString)
                    if attrib.TagString == "CIRCUITO":
                        attrib.TextString = all_handles_with_correct_circuit_num[entity.Handle]

    # Close the document
    # doc.Save()
    # doc.Close()
    print("AutoCAD file circuit numbers updated successfully.")


def browse_file():
    filename = filedialog.askopenfilename()
    autocad_file_entry.delete(0, 'end')
    autocad_file_entry.insert(0, filename)


######## UI #########################

root = Tk()
root.title("Dinukas Program")

window_width = 600
window_height = 300

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

position_top = int(screen_height / 2 - window_height / 2)
position_right = int(screen_width / 2 - window_width / 2)

root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")

label2 = Label(root, text='AutoCAD File', font=("Arial bold", 12))
label2.pack(padx=5, pady=20)
autocad_file_entry = Entry(root, width=50, font=("Arial bold", 12))
autocad_file_entry.pack(padx=5, pady=10)
browse_btn = Button(root, text='Browse', font=("Arial bold", 12), command=browse_file, bg='green', fg='white', padx=25,
                    pady=5)
browse_btn.pack(padx=5, pady=10)

frame = Frame(root)
frame.pack(padx=5, pady=10)

button_start = Button(frame, text='Start Balancing', font=("Arial bold", 13), bg='blue', fg='white', padx=25, pady=5,
                      command=startBalancing)
button_start.pack(padx=5, pady=20)

root.mainloop()
