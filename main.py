from Phase2PanelBoard import balance2PhasePanelBoard
from Phase3PanelBoard import balance3PhasePanelBoard
from openpyxl.utils import get_column_letter
from collections import Counter
from collections import Counter
from openpyxl.styles import Border, Side, Alignment, Font

from openpyxl import Workbook
import warnings

autocad_file_path = "C:\\Users\\sachi\\OneDrive\\Documents\\Dinukas Program\\output.txt"
input_filename = "generated_input_file.xlsx"
output_filename = "temp_output.xlsx"


def mergeCells(sheet, cell1, cell2):
    start_cell = f"{get_column_letter(cell1[1])}{cell1[0]}"
    end_cell = f"{get_column_letter(cell2[1])}{cell2[0]}"
    sheet.merge_cells(f"{start_cell}:{end_cell}")


def readInputFileFromAutocad(filename):
    print("Start Reading generated file from AutoCAD")
    load_details_dict = {}
    circuit_details_dict = {}
    f = open(filename)
    lines = f.readlines()

    # Read panel board name, number of phases and number of circuits from autocad file
    panel_board_name = lines[0].split(":")[1].strip()
    num_of_phases = int(lines[1].split(":")[1].strip())
    max_circuits = int(lines[2].split(":")[1].strip())

    # Read line by line
    for line in lines[3:]:
        line = line.strip()

        data = line.split()
        circuit_number = data[0].split(":")[1].strip()
        object_id = data[1].split(":")[1].strip()
        poles = int(data[2].split(":")[1].strip())
        power = float(data[3].split(":")[1].strip())
        load_number = int(data[5].split(":")[1].strip())

        # Add each unique load to load_details_dict
        # Each load can be accessed by load number

        if load_number not in load_details_dict.keys():
            load_details_dict[load_number] = {"object_id": object_id, "poles": poles, "power": power}

        # Update circuit details dict
        temp = []
        if circuit_number in circuit_details_dict.keys():
            temp = circuit_details_dict[circuit_number]
        temp.append(load_number)
        circuit_details_dict[circuit_number] = temp

    sorted_load_numbers = sorted(load_details_dict.keys())
    for circuit_num in circuit_details_dict.keys():
        counter = Counter(circuit_details_dict[circuit_num])
        circuit_details_dict[circuit_num] = counter

    print("Read generated file from AutoCAD")
    return load_details_dict, sorted_load_numbers, circuit_details_dict, panel_board_name, num_of_phases, max_circuits


if __name__ == '__main__':

    load_details_dict, sorted_load_numbers, circuit_details_dict, panel_board_name, num_of_phases, max_circuits = readInputFileFromAutocad(
        autocad_file_path)

    print("Start generating input excel file")
    workbook = Workbook()
    worksheet = workbook.active

    mergeCells(worksheet, (1, 1), (2, 1))
    worksheet.cell(row=1, column=1).value = "CIRCUIT LIST"

    for index, load_num in enumerate(sorted_load_numbers):
        worksheet.cell(row=1, column=index + 2).value = "LOAD #" + str(load_num)
        worksheet.cell(row=2, column=index + 2).value = load_details_dict[load_num]["power"]

    worksheet.cell(row=1, column=len(sorted_load_numbers) + 2).value = "Total power"
    worksheet.cell(row=2, column=len(sorted_load_numbers) + 2).value = "(WATTS)"

    for index, circuit_num in enumerate(circuit_details_dict.keys()):
        worksheet.cell(row=3 + index, column=1).value = "CIRCUIT #" + str(circuit_num)
        loads = circuit_details_dict[circuit_num]

        circuit_power = 0
        for load_num, count in loads.items():
            load_index = sorted_load_numbers.index(load_num)
            worksheet.cell(row=3 + index, column=2 + load_index).value = count
            circuit_power += load_details_dict[load_num]["power"] * count

        worksheet.cell(row=3 + index, column=2 + len(sorted_load_numbers)).value = circuit_power

        worksheet.cell(row=3 + index, column=1).value = "CIRCUIT #" + str(circuit_num)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in range(1, len(circuit_details_dict.keys()) + 3):
        for col in range(1, len(sorted_load_numbers) + 3):
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).border = thin_border
            worksheet.cell(row=row, column=col).font = Font(size=10)

    # Add fixed width to each column
    for col in range(1, len(circuit_details_dict.keys()) + 3):
        width = 15
        if col == 1:
            width = 25
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = width

    workbook.save('generated_input_file.xlsx')
    print("Completed generating input excel file")

    if num_of_phases == 3:
        balance3PhasePanelBoard(input_filename, output_filename, max_circuits)

    else:
        balance2PhasePanelBoard(input_filename, output_filename, max_circuits)
